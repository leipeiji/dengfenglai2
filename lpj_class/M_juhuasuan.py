# -*- coding: utf-8 -*-
from ..config import *
from ..M_functionToolKit import *
import requests
import re


from openpyxl import Workbook
from openpyxl import load_workbook
# print('######################################\n')
#http://www.ahatao.com/listname/?lid=556605256954
import pandas as pd
#得到没有解析的产品信息 json格式
# 年货节 活动是参考 聚划算地址
#  https://ju.taobao.com/json/jusp2/ajaxGetTpFloor.json?_ksTS=1516753995784_690&callback=_mtp_J_dyp2s340nfs&urlKey=o/nianhuo&floorIndex=7&pc=true&impid=fODMMU7wxKo&ext=page:3;
# _ksTS
# 1516754024026_788
# callback
# _mtp_J_dyp2s340nfs
# ext
# page:4;  就 这个是变量
# floorIndex	 7
# impid
# fODMMU7wxKo
# pc
# true
# urlKey
# o/nianhuo

MYCOUNT=0
class JuhuasuanSingle(object):
    def __init__(self,save_dir=r'{}:\聚划算'.format(ROOT_DIR), is_save_database=1,keyWord = '聚划算全部产品_',
                 temp_keyWord = '',is_save_pic = 1,totalPage=None):
        self.MYCOUNT=0
        self.is_save_database = is_save_database
        self.database, self.tableName = 'taobao', 'taobao_juhuasuan_single_product'
        self.database_field_list = ['page', 'itemId', 'crawlDate', 'crawlTime', 'startTime', 'endTime', 'totalStock',
                               'price',
                               'soldCount', 'remindNum', 'category', 'juItemUrl', 'picUrl', 'itemUrl', 'sellingPoint1',
                               'sellingPoint2', 'sellingPoint3', 'operator']

        self.save_dir = save_dir
        self.temp_keyWord = temp_keyWord  # 在大活动时，每天保存多次数据，用的比较多 这里区分下目录 和文件名
        self.keyWord = keyWord + self.temp_keyWord
        self.tempKeyWord = '每页数据_' + self.temp_keyWord
          # 建立目录用
        self.firstPath, self.TempPath=self.dirFun()
        self.is_save_pic = is_save_pic
        self.startPage = 1
        self.totalPage = totalPage if totalPage is not None else self.getAllJuListAjax(p=1, is_totalPage=True)  # 总页码会自动识别

    def getAllJuListAjax(self,p=1,is_totalPage=False, RequestsUrl='https://ju.taobao.com/json/tg/ajaxGetHomeItemsV2.json'):

        try:
              query = {'page': p,'type':0,'timeFilter':'todayall','stype':'soldCount'}
              headers = {'user-agent': random.choice(user_agent_list)}

              #timeFilter= (today | old  | foreshow |todayall) 分别表示 今日新开团 昨日开团 预告 所有的
              #'stype': 'soldCount' 按销量排序
              #stype=activityPrice&reverse=up 价格从低到高排序
              r = requests.get(RequestsUrl, params=query, headers=headers)
              print(r.status_code,r.url)
              if r.status_code==200:
                  c = r.text
                  new_json = json.loads(c)
                  # pprint(new_json.keys())
                  # pprint(new_json)
                  if is_totalPage:
                      allPage=new_json.get('totalPage',1)
                      return int(allPage)
                  if len(new_json['itemList'])==0:
                      myFormat('%d页\t没有数据了'%p)
                      return None
                  self.MYCOUNT = 0
                  return new_json['itemList']
              else:
                  print('服务器拒绝')
                  return None
        except Exception as e:
            self.MYCOUNT += 1
            if self.MYCOUNT < 3:
                text='getAllJuListAjax,第%d次链接出现问题，正在请求第%d次链接......' % (self.MYCOUNT, self.MYCOUNT + 1)
                my_logging(text,filename='M_allJuHuaSuan.log')
                print('等待%d秒.....' % (2 * self.MYCOUNT))
                time.sleep(2 * self.MYCOUNT)
                print(e, text)
                print('错误时的页码是', p)
                return self.getAllJuListAjax(p=p,is_totalPage=is_totalPage ,RequestsUrl=RequestsUrl)
            else:
                self.MYCOUNT = 0
                print('error stop')
                return None


    #获取每一页的所有产品
    def getEveryJuInfo(self,allList,p=1):
        if allList==None or len(allList)<1:
            print('%d页没有获取到数据,可能是超出了页码，或本页出现错误了'%p)
            return None
        allProduct = []
        # AllExcelHead = ['页码', '产品ID', '抓取时间', '开抢时间', '结束时间', '备货量', '参聚价格', '销量', '聚划算链接', '聚划算主图', '单品链接', '卖点1', '卖点2','卖点3']
        # 用csv文件print()
        try:
            for k, js in enumerate( allList):

                picUrlNew = js['baseinfo'].get('picUrlNew')
                if not picUrlNew:
                    picUrlNew = js['baseinfo'].get('picUrl')
                if not 'http' in picUrlNew:
                    picUrlNew = 'https:' + picUrlNew

                tbFirstCatId=js['baseinfo'].get('tbFirstCatId')
                try:
                    price=str(js['price'].get('actPrice', 0))
                    if '?' in price: # 过滤猜价格的活动时
                        price=0
                except:
                    price=0


                innerProduct = [p,
                                str(js['baseinfo'].get('itemId', '')),
                                time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                                js['baseinfo'].get('ostimeText', ''), js['baseinfo'].get('leftTime', ''),
                                js['baseinfo'].get('totalStock', 0),
                                price, js['remind'].get('soldCount', 0),
                                js['remind'].get('remindNum', 0),taobao_all_category_reversed.get(tbFirstCatId),
                                'https:' + js['baseinfo'].get('itemUrl', '#'), picUrlNew,
                                'https://item.taobao.com/item.htm?id=' + str(js['baseinfo'].get('itemId', '')),
                                str(js['merit'].get('down', '')), str(js['merit'].get('up', '')),
                                str(js['merit'].get('desc', '')),
                                ]
                if k%10==0:
                    print(innerProduct)
                allProduct.append(innerProduct)
            return allProduct
        except Exception as e :
            print(e)
            myFormat('获取第%d页产品出现错误' % p,symbol='☹')
            return None

    #存到excel文件 这个用在excel 2007及以上的版本
    def EveryPageWriteExcel2016(self,ContainPicUrlList,p=1,sheetName='data'):
        if ContainPicUrlList==None or len(ContainPicUrlList)==0:
            print('%d页\t没有获取到数据,请检查原始数据元列表'%p)
            return None
        # doc = r'{}\{}.{}'.format(path, filename, 'xlsx')
        AllExcelHead = ['0页码', '1产品ID', '2抓取时间', '3开抢时间', '4结束时间', '5备货量', '6参聚价格', '7销量',
                        '8开团提醒数', '9所属类目',
                        '10聚划算链接',
                        '11聚划算主图', '12单品链接', '13卖点1', '14卖点2', '15卖点3']
        # print(TempPath)
        doc = r'{}\{}.{}'.format(self.TempPath,str(p), 'xlsx')
        # print(doc)
        try:
            # 在内存创建一个工作簿obj
            wb = Workbook()
            ws=wb.active
            #给sheet明个名
            ws.title = sheetName
            # 向第一个sheet页写数据吧 格式 ws2['B1'] = 4
            ws.append(AllExcelHead)
            k = 0
            for line in ContainPicUrlList:
                try:
                    #line是一个列表 类似[1,2,3.4.5]
                    ws.append(line)
                    k += 1
                    if k%10==0:
                        print('写入第%d条记录完毕' % (k))
                except:
                    print('第%d条记录有问题，已经忽略' % k)
                    continue
            else:
                print('###############恭喜你，第%d页写入完毕#####################'%p)
                # i += 1
            # 工作簿保存到磁盘
            wb.save(doc)
        except:
            print('出现问题了')
            pass

    #合并多个文件 成一个文件
    def combineEveryPageInfoToOne(self,):
        readDirFile=['{}\{}.{}'.format(self.TempPath,x,'xlsx') for x in range(1,self.totalPage+1)]
          # 定义一个存放原来整体数据list
        AllInfoList=[]
        count=1
        for doc in readDirFile:
            if os.path.isfile(doc):
                wb1 = load_workbook(filename=doc)
                sheet_ranges = wb1.active
                # 分别获取原来数据的 最后一行和最后一列的值
                endRowNum = sheet_ranges.max_row
                endColNum = sheet_ranges.max_column
                print(endRowNum,endColNum,'\t最大行列')
                # test=sheet_ranges['A1'].value
                outerValueList = []
                for m in range(2, endRowNum + 1):
                    innerValueList = []  # 定义个一个内部存储数据的list，存放表格中每一行
                    for v in range(1, endColNum + 1):
                        ColNum = '%c' % (64 + v)  # 用ascii转化下，这里用数字转化成字母 比如65是A，66是B 这样获取 表格的 A B ...列
                        cellValue = sheet_ranges[ColNum + str(m)].value
                        # print(ColNum)
                        innerValueList.append(cellValue)
                    outerValueList.append(innerValueList)

                AllInfoList+=outerValueList
                print('第%d个文件处理完毕了'%count)
                count+=1
        else:
            print('全部%d个文件合并完毕'%count)
        time.sleep(1)
        # 写入上面合并后的数据到一个总文件中
        allJuProductPath = '{}\{}\{}.{}'.format(self.save_dir, time.strftime("%Y-%m-%d", time.localtime()),
                                                    time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime()) + self.keyWord, 'xlsx')
        sheetTitle=os.path.basename(allJuProductPath)
        try:
            # 在内存创建一个工作簿obj
            wb6 = Workbook()
            ws = wb6.active
            # 给sheet明个名
            ws.title = sheetTitle
            # 向第一个sheet页写数据吧 格式 ws2['B1'] = 4
            AllExcelHead = ['0页码', '1产品ID', '2抓取时间', '3开抢时间', '4结束时间', '5备货量', '6参聚价格',
                            '7销量', '8开团提醒数', '9所属类目',
                            '10聚划算链接',
                            '11聚划算主图', '12单品链接', '13卖点1', '14卖点2', '15卖点3']
            ws.append(AllExcelHead)
            k = 0
            for line in AllInfoList:
                try:
                    # line是一个列表 类似[1,2,3.4.5]
                    ws.append(line)
                    k += 1
                    print('写入第%d条记录完毕' % (k))
                except:
                    print('第%d条记录有问题，已经忽略' % k)
                    continue
            if k == len(AllInfoList):
                print('###############恭喜你，全部记录写入完毕#####################')
            # 工作簿保存到磁盘
            wb6.save(allJuProductPath)
            wb6.close()
        except:
            print('出现问题了')
            pass

    # 用pandas 来合并多个excel ，效率更高
    def combineEveryPageInfoToOneV2(self,):
        try:
            readDirFile = ['{}\{}.{}'.format(self.TempPath, x, 'xlsx') for x in range(1, self.totalPage + 1)]
            dataList=[]
            for doc in readDirFile:
                try:
                    if os.path.isfile(doc):
                        data = pd.read_excel(doc, sheetname='data')
                        dataList.append(data)
                except Exception as e:
                    print(e,'合并\t%s\t文档出错,已经跳过'%doc)
                    continue
            dataAll = pd.concat(dataList)
            dataAll = dataAll.drop_duplicates(['1产品ID']) # 数据去重 ，因为聚划算聚数据 每页的产品不停的在变化，
            # 一个产品可能同时出现在 2页，所以需要去重
            # print('合并后的类型是',type(dataAll))

            allJuProductPath = '{}\{}\{}.{}'.format(self.save_dir, time.strftime("%Y-%m-%d", time.localtime()),
                                                    time.strftime("%Y-%m-%d_%H-%M-%S", time.localtime()) + self.keyWord, 'xlsx')
            # sheetTitle = os.path.basename(allJuProductPath)

            dataAll.to_excel(allJuProductPath, index=False, sheet_name='当天合并数据')
            myFormat('保存excel成功,路径是 %s'%allJuProductPath)
            if self.is_save_database:
                df=dataAll
                get_date = df['2抓取时间'].map(lambda x: x.split(' ')[0])
                get_time = df['2抓取时间'].map(lambda x: x.split(' ')[1])
                del df['2抓取时间']  # 删除 抓取时间这列
                df.insert(2, 'get_date', get_date)  # 插入新列，第一个参数，插入的位置，第二个参数，新列的名字，第三个参数 ，一个列表值list
                df.insert(3, 'get_time', get_time)
                df['operator']=OPERATOR_NAME
                df.columns=self.database_field_list
                con_database_from_pd(self.database, self.tableName, df)
                myFormat('数据库名【%s】,表名是【%s】'%(self.database, self.tableName))

        except Exception as e:
            print(e,'合并数据出了问题,检查 函数 combineEveryPageInfoToOneV2')
            return



    #下载图片
    def downloadPic(self,ContainPicUrlList,p=1):
        if ContainPicUrlList==None or len(ContainPicUrlList)==0:
            print('%d页\t没有获取到数据,请检查原始数据元列表'%p)
            return None
        try:
            k=0
            for n in ContainPicUrlList:
                # time.sleep(1)
                k += 1
                PrePic =  '页码_' + str(p)+ '_排名_' + str(k) + '_价格_' + str(n[6]).replace('?','X') +'_销量_'+\
                          str(n[7]).replace('/','')+'_开团提醒_'+ str(n[8])+ '_备货_' + str(n[5])+ '_' + str(n[1])
                file_path = '{0}\{1}.{2}'.format(self.firstPath, PrePic, 'jpg')
                #print(file_path,n[-2])
                try:
                    print(n[11])
                    response = requests.get(n[11], headers=headers)
                    with open(file_path, 'wb') as f:
                        f.write(response.content)
                        f.close()
                    if k%10==0:
                        print('保存第%d张图片完毕了' %(k))
                    # if k==len(ContainPicUrlList):
                except Exception as e:
                    print(e,'保存图片---%s---时出错了，本张图片没有保存' %(PrePic))
                    continue
            else:
                print('##################恭喜你，第%d页图片全部保存完毕!#######################' % p)
        except Exception as e:
            print(e,'保存图出错')
            pass

    def dirFun(self,):
        try:
            firstPath = '{}\{}\{}'.format(self.save_dir, time.strftime("%Y-%m-%d", time.localtime()), self.keyWord)
            if not os.path.exists(firstPath):
                os.makedirs(firstPath)
            TempPath = '{}\{}\{}'.format(self.save_dir, time.strftime("%Y-%m-%d", time.localtime()), self.tempKeyWord)
            if not os.path.exists(TempPath):
                os.makedirs(TempPath)
            return firstPath,TempPath

        except Exception as e:
            print(e,'建立目录函数出错了,dirFun')

    #主运行程序
    def main(self,p):
        myFormat('正在处理第%d页'%p,symbol='.',allLength=150,fillMode='right')
        curPageAjaxInfo=self.getAllJuListAjax(p=p)
        everyPageInfoList=self.getEveryJuInfo(curPageAjaxInfo,p=p) # 获取每页数据
        self.EveryPageWriteExcel2016(everyPageInfoList,p=p) # 下载图片时，这个必须打开 因为这里面改变了 everyPageInfoList的值
        if self.is_save_pic:
            self.downloadPic(everyPageInfoList, p=p)
        myFormat('第%d页处理完毕'%p,symbol='#',allLength=150)


# ╭~~~╮
# (o^.^o)
# -----------功能---------  聚划算全部产品  ---------------更新时间  2017-11-05 ---------
# ----------聚划算单品保存，可以同步下载图片 ，保存到数据库，最终存储的 数据示例如下
# title_list=['0页码', '1产品ID', '2抓取时间', '3开抢时间', '4结束时间', '5备货量', '6参聚价格', '7销量', '8开团提醒数', '9所属类目',
#                     '10聚划算链接','11聚划算主图', '12单品链接', '13卖点1', '14卖点2', '15卖点3']
# data_list=[4, '522839448042', '2017-11-20 08:53:59', '11月17日 10:00准时开抢', '4分钟', 8000, '118', 882, 317, '女装/女士精品',
#                    'https://detail.ju.taobao.com/home.htm?id=10000060638011&item_id=522839448042', 'https://gju3.alicdn.com/tps/i2/1776921957/TB2bhFDdh6I8KJjSszfXXaZVXXa_!!0-juitemmedia.jpg', 'https://item.taobao.com/item.htm?id=522839448042', "['退货无忧15天包退']", "['今冬热销爆款', '100%棉质面料', '前300名减20']", "['热销3万多件，万千妈妈的选择！', '淘宝官方质检报告：含棉100%', '孝心季节，冬季给妈妈买一件好的贴心的保暖棉衣！']"]写入第21条记录完毕
# ╭~~~╮
# (o^.^o)







